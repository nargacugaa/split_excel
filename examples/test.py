import os
import openpyxl
from openpyxl import Workbook
from datetime import datetime, timedelta

def excel_serial_date_to_date(serial):
    # Excel dates are based on 1900-01-01, but Excel incorrectly treats 1900 as a leap year
    base_date = datetime(1899, 12, 30)
    delta = timedelta(days=serial)
    return (base_date + delta).strftime('%Y/%m/%d')

def split_excel_file(input_file, output_prefix, rows_per_file):
    workbook = openpyxl.load_workbook(input_file, data_only=True)
    if 'Sheet1' not in workbook.sheetnames:
        raise ValueError("Cannot find 'Sheet1'")
    
    sheet = workbook['Sheet1']
    current_row = 2  # Start from the second row (first row is header)
    file_index = 1

    # Extract the first row data (header)
    first_row = list(sheet.iter_rows(min_row=1, max_row=1, values_only=True))[0]

    while current_row <= sheet.max_row:
        new_workbook = Workbook()
        new_sheet = new_workbook.active

        # Write the first row data (header)
        for col, cell_value in enumerate(first_row, start=1):
            new_sheet.cell(row=1, column=col, value=cell_value)

        for row in sheet.iter_rows(min_row=current_row, max_row=min(current_row + rows_per_file - 1, sheet.max_row), values_only=True):
            for col, cell_value in enumerate(row, start=1):
                if isinstance(cell_value, (int, float)):
                    new_sheet.cell(row=current_row - (current_row // rows_per_file * rows_per_file) + 1, column=col, value=cell_value)
                elif isinstance(cell_value, str):
                    new_sheet.cell(row=current_row - (current_row // rows_per_file * rows_per_file) + 1, column=col, value=cell_value)
                elif isinstance(cell_value, bool):
                    new_sheet.cell(row=current_row - (current_row // rows_per_file * rows_per_file) + 1, column=col, value='TRUE' if cell_value else 'FALSE')
                elif isinstance(cell_value, datetime):
                    new_sheet.cell(row=current_row - (current_row // rows_per_file * rows_per_file) + 1, column=col, value=cell_value.strftime('%Y/%m/%d'))
                else:
                    new_sheet.cell(row=current_row - (current_row // rows_per_file * rows_per_file) + 1, column=col, value='')

            current_row += 1

        output_file = f"{output_prefix}_{file_index:03}.xlsx"
        new_workbook.save(output_file)

        print(f"已拆分为：{file_index} 个Excel")

        file_index += 1

def main():
    current_dir = os.getcwd()
    input_file = None

    # 查找当前目录中的第一个 .xlsx 或 .xls 文件
    for file_name in os.listdir(current_dir):
        if file_name.endswith('.xlsx') or file_name.endswith('.xls'):
            input_file = file_name
            break
    
    if input_file is None:
        raise FileNotFoundError("No .xlsx or .xls file found in the current directory")

    # 创建 result 目录
    output_dir = os.path.join(current_dir, 'result')
    os.makedirs(output_dir, exist_ok=True)

    # 设置输出文件前缀
    output_prefix = os.path.join(output_dir, 'output')

    rows_per_file = 10000

    split_excel_file(input_file, output_prefix, rows_per_file)

    # 完成
    print("完成！！ Excel 文件已成功拆分 \r\n按回车键关闭窗口")
    input()

if __name__ == "__main__":
    main()